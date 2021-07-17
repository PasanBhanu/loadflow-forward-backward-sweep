import numpy
import math
import time
from openpyxl import load_workbook
from Functions import *
from LineModels import *
from TransformerModels import *
from DG import *
from LoadModels import *

# Global Variables - Symmetrical Components
a = complex(-0.5, math.sqrt(3)/2)
a_2 = complex(-0.5, -1 * math.sqrt(3)/2)
nodeThreePhaseVoltagePU = numpy.array([[1],[a_2],[a]])
converged = False

# Time
startTime = time.time()

# Load data from excel sheet
wb = load_workbook(filename = 'dat.xlsx')
edgesSheet = wb['edges']
nodesSheet = wb['nodes']
voltagesSheet = wb['voltages']
settingsSheet = wb['settings']
sweepOutputSheet = wb['fbsweep']
outputSheet = wb['finaldata']

# Add data to arrays
edges = []
for value in edgesSheet.iter_rows(min_row=2,min_col=1,max_col=5,values_only=True):
    # Calculate Line Impedances
    if value[3] == "L":
        epm = value[4].split(",")
        # SIGN => (resistance_p, gmr_p, resistance_n, gmr_n, phases, length, isNeutralAvailable, frequency, soilResistivity, type, l_12, l_13, l_23, l_1n, l_2n, l_3n)
        lineParameters = calculateLineResistance(float(epm[0]),float(epm[1]),float(epm[2]),float(epm[3]),int(epm[4]),float(epm[5]),bool(epm[6]),float(epm[7]),float(epm[8]),epm[9],float(epm[10]),float(epm[11]),float(epm[12]),float(epm[13]),float(epm[14]),float(epm[15]))
        line = (value[0], lineParameters, value[2], value[3])
        edges.append(line)
    else:
        edges.append(value)

nodes = []
for value in nodesSheet.iter_rows(min_row=2,min_col=1,max_col=4,values_only=True):
    nodes.append(value)

voltages = []
for value in voltagesSheet.iter_rows(min_row=2,min_col=1,max_col=3,values_only=True):
    voltages.append(value)

settings = []
for value in settingsSheet.iter_rows(min_row=1,min_col=1,max_col=2,values_only=True):
    settings.append(value)

# Network Limits
startNode = 1
lastNode = maxValueFromExcelColumns(edges, 2)

# Clean Workbook
wb.remove(sweepOutputSheet)
wb.remove(outputSheet)
sweepOutputSheet = wb.create_sheet('fbsweep')
outputSheet = wb.create_sheet('finaldata')
sweepOutputSheet['A1'] = 'Node'
outputSheet['A1'] = 'Node'
for row in range(2,lastNode + 2):
    sweepOutputSheet.cell(column=1, row=row, value=row-1)
    outputSheet.cell(column=1, row=row, value=row-1)

# Calculate Bases
bases = []
# Voltages for Each Node (Initialisation on Calculated Voltages Array)
calculatedVoltages = []

rating = searchArray(voltages, 1)[2]
# Get polar value and converted to cartesian
tmpPolarVoltage = searchArray(voltages, 1)[1]
voltage = convertToCartisan(float(tmpPolarVoltage.split(",")[0]), float(tmpPolarVoltage.split(",")[1]))

def calculateBase(startNode, voltage, rating, connection):
    connectedNodes = getConnectedNodes_BS(edges, startNode)
    for connectedNode in connectedNodes:
        if connectedNode[3] == 'T':
            transformerData = searchDoubleArray(edges, startNode, connectedNode[2])[4].split(",")
            rating = float(transformerData[0])
            voltage = complex(transformerData[2])
            if transformerData[5] == 'YD':
                connection = 'D'
            else:
                connection = 'S'
        else:
            if connection == 'S':
                voltage = voltage * math.sqrt(3)
        bases.append([connectedNode[2], voltage, rating])
        calculateBase(connectedNode[2], voltage, rating, connection)
        # Set Voltage for Each Node
        calculatedVoltages.append([connectedNode[2], nodeThreePhaseVoltagePU])

bases.append([1, voltage, rating])
calculatedVoltages.append([1, nodeThreePhaseVoltagePU])
# Calculate base assuming initial connection is Star (Need developments)
calculateBase(1, voltage, rating, 'S')

# Network iteration
iteration = 1
tolerance = settings[0][1]
maxIterations = settings[1][1]
roundFactor = settings[2][1]
upperLimitViolation = settings[3][1]
lowerLimitViolation = settings[4][1]

# Comparison
calculatedVoltagesOfPreviousIteration = []

### LOOP ###

while (iteration <= maxIterations and converged == False):

    # Variables for Sweep
    calculatedCurrents = []
    calculatedEdgeCurrents = []

    calculatedTotalCurrentsAtNode = [] # Used for algorithmic purpose only

    # Calculate Node Currents
    for node in nodes:
        nodeId = node[0]
        nodeConnection = node[1]
        nodeType = node[2]
        voltage = searchArray(calculatedVoltages, nodeId)[1]
        
        nodeCurrent = numpy.array([[complex(0,0)],[complex(0,0)],[complex(0,0)]])
        if nodeType == 'L':
            # Power (kVA), CPQ, CC, CI
            loadParameters = node[3].split(",")
            base = searchArray(bases, nodeId)
            pu_Power = numpy.array([[complex(loadParameters[0])/base[2]],[complex(loadParameters[1])/base[2]],[complex(loadParameters[2])/base[2]]])

            if nodeConnection == 'S':
                # LN Voltage is Required
                nodeCurrent = loadCurrentStar(voltage, pu_Power, float(loadParameters[3]), float(loadParameters[4]), float(loadParameters[5]))
            else:
                nodeCurrent = loadCurrentDelta(voltage, pu_Power, float(loadParameters[3]), float(loadParameters[4]), float(loadParameters[5]))

        elif nodeType == 'C':
            # kV, kVar
            loadParameters = node[3].split(",")
            base = searchArray(bases, nodeId)
            # Zbase = V_2 / kVA
            z_base = numpy.power((base[1]),2) / (base[2] * 1000)
            kV = numpy.array([[complex(loadParameters[0])],[complex(loadParameters[1])],[complex(loadParameters[2])]])
            kVar = numpy.array([[complex(loadParameters[3])],[complex(loadParameters[4])],[complex(loadParameters[5])]])
            if nodeConnection == 'S':
                nodeCurrent = shuntCapacitanceStar(voltage, kV, kVar, z_base)
            else:
                nodeCurrent = shuntCapacitanceDelta(voltage, kV, kVar, z_base)
        elif nodeType == 'D':
            loadParameters = node[3].split(",")
            base = searchArray(bases, nodeId)
            pu_Power = numpy.array([[complex(loadParameters[0])/base[2]],[complex(loadParameters[1])/base[2]],[complex(loadParameters[2])/base[2]]])
            if nodeConnection == 'S':
                nodeCurrent = generatorStar(voltage, pu_Power)
            else:
                nodeCurrent = generatorStarDelta(voltage, pu_Power)
        
        # Add Total Nodal Current to Array
        totalNodeCurrent = searchArray(calculatedCurrents, nodeId)
        if (totalNodeCurrent == 0):
            calculatedCurrents.append([nodeId, nodeCurrent])
        else:
            calculatedCurrents.remove(totalNodeCurrent)
            calculatedCurrents.append([nodeId, numpy.add(totalNodeCurrent[1],nodeCurrent)])

    # Backward Sweep
    for i in range(lastNode,0, -1):
        connectedNodes = getConnectedNodes_BS(edges, i)
        destinationNode = getDestinationNode_BS(edges, i)

        # SIGN => (Node ID, Current Matrix)
        current = searchArray(calculatedCurrents,i)
        if current == 0:
            totalCurrentAtNode = numpy.array([[complex(0,0)],[complex(0,0)],[complex(0,0)]])
        else:
            totalCurrentAtNode = current[1]

        for connectedNode in connectedNodes:
            if (connectedNode[3] == 'L'):
                # SIGN => (Node ID, Current Matrix)
                connectedNodeTotalCurrent = searchArray(calculatedTotalCurrentsAtNode,connectedNode[2])
                totalCurrentAtNode = numpy.add(totalCurrentAtNode, connectedNodeTotalCurrent[1])
            else:
                # Do TF Convert
                # SIGN => (Rating, Primary Voltage, Secondary Voltage, Z, XR, TF Type, Angle)
                tfParameters = connectedNode[4].split(",")
                angle = (2 * math.pi * float(tfParameters[6])) / 360
                connectedNodeTotalCurrent = searchArray(calculatedTotalCurrentsAtNode,connectedNode[2])
                voltage = searchArray(calculatedVoltages, i)[1]

                # Select respective function based on TF type
                if (tfParameters[5] == 'YD'):
                    convertedCurrentAtThisNode = backwardTfYgD(voltage, connectedNodeTotalCurrent[1], float(tfParameters[3])/100, angle)
                elif (tfParameters[5] == 'YY'):
                    connectedNodeTotalCurrent = searchArray(calculatedTotalCurrentsAtNode,connectedNode[2])
                    convertedCurrentAtThisNode = numpy.add(totalCurrentAtNode, connectedNodeTotalCurrent[1])
                else:
                    convertedCurrentAtThisNode = 0
                totalCurrentAtNode = numpy.add(totalCurrentAtNode, convertedCurrentAtThisNode)
            
        calculatedTotalCurrentsAtNode.append([i, totalCurrentAtNode])

        if destinationNode != 0 and destinationNode[3] == 'L':
            calculatedEdgeCurrents.append([destinationNode[0], totalCurrentAtNode, i])
        elif destinationNode != 0 and destinationNode[3] == 'T':
            tfParameters = destinationNode[4].split(",")
            if tfParameters[5] == 'YY':
                calculatedEdgeCurrents.append([destinationNode[0], totalCurrentAtNode, i])
    
    # Clear Voltages Array
    calculatedVoltages = []
    
    # Forward Sweep
    for i in range(1, lastNode + 1, 1):
        edgeData = getPreviousNode_FS(edges, i)
        if (edgeData == 0):
            newVoltage = nodeThreePhaseVoltagePU
        else:
            if (edgeData[3] == 'L'):
                # SIGN => (Node ID, Voltage Matrix)
                previousNodeVoltage = searchArray(calculatedVoltages,edgeData[0])
                # SIGN => (Start Node ID, Current Matrix, End Node ID)
                edgeCurrent = searchDoubleArray(calculatedEdgeCurrents, edgeData[0], i)
                # Convert Impedance to PU
                base = searchArray(bases, i)
                # Zbase = V_2 / kVA
                z_base = numpy.power((base[1]),2) / (base[2] * 1000)
                # Zero Sequence Impedance, Positive Sequence Impedance
                edgeParameters = edgeData[1].split(",")
                newVoltage = approximateLineModel(previousNodeVoltage[1], edgeCurrent[1] , complex(edgeParameters[0])/z_base, complex(edgeParameters[1])/z_base)
            else:
                # SIGN => (Node ID, Voltage Matrix)
                previousNodeVoltage = searchArray(calculatedVoltages,edgeData[0])

                nodeTotalCurrent = searchArray(calculatedTotalCurrentsAtNode,i)

                nodePreviousIterationVoltage = searchArray(calculatedVoltagesOfPreviousIteration,i)
                if nodePreviousIterationVoltage == 0:
                    nodePreviousIterationVoltage = nodeThreePhaseVoltagePU
                else:
                    nodePreviousIterationVoltage = nodePreviousIterationVoltage[1]

                tfParameters = edgeData[4].split(",")
                angle = (2 * math.pi * float(tfParameters[6])) / 360

                # Select respective function based on TF type
                newVoltage = fowardTfYgD(previousNodeVoltage[1], nodePreviousIterationVoltage, nodeTotalCurrent[1], float(tfParameters[3])/100, angle)


        calculatedVoltages.append([i, newVoltage])


    # Output results variables
    outputPolarStr = ''
    outputPUStr = ''
    lowerViolationArray = []
    upperViolationArray = []

    ## Add data to excel
    sweepOutputSheet.cell(column=iteration + 1, row=1, value='Sweep ' + str(iteration))
    for row in range(2,lastNode + 2):
        printVoltage = searchArray(calculatedVoltages,row-1)[1]
        base = searchArray(bases, row-1)
        polarVoltages = []
        puVoltages = []

        # Limit Checking Variables
        isUpperViolated = isLowerViolated = False
        allowedUpperLimit = base[1] * upperLimitViolation
        allowedLowerLimit = base[1] * lowerLimitViolation

        for complexPUVoltage in printVoltage:
            # Limit Check Calculations
            basePolar = getPolarMagnitude(base[1],roundFactor)
            voltagePolar = getPolarMagnitude(complexPUVoltage[0] * base[1],roundFactor)
            
            if (basePolar - voltagePolar) > 0:
                # Lower Limit
                if (basePolar - voltagePolar) >= allowedLowerLimit:
                    isLowerViolated = True
            else:
                # Upper Limit
                if (voltagePolar - basePolar) >= allowedUpperLimit:
                    isUpperViolated = True


            polarVoltages.append(convertToPolar(complexPUVoltage[0] * base[1], roundFactor))
            puVoltages.append(str(complexPUVoltage[0]))

        # Print Violations
        if isLowerViolated:
            lowerViolationArray.append((row-1, polarVoltages)) 

        if isUpperViolated:
            upperViolationArray.append((row-1, polarVoltages)) 
        
        outputPolarStr += '\n' + str(row-1) + '\t' + "\t\t".join(polarVoltages)

        sweepOutputSheet.cell(column=iteration + 1, row=row, value=str(", ".join(polarVoltages)))

    wb.save(filename = 'dat.xlsx')

    # Check Error
    if iteration > 1:
        totalMaximumError = 0
        for i in range(1, lastNode + 1, 1):
            lastIterationVoltage = searchArray(calculatedVoltagesOfPreviousIteration,i)[1]
            newIterationVoltage = searchArray(calculatedVoltages,i)[1]
            base = searchArray(bases, i)
            errorMatrixPU = numpy.absolute(numpy.subtract(lastIterationVoltage, newIterationVoltage))
            errorMatrix = numpy.multiply(errorMatrixPU, base[1])
            maximumError = numpy.max(errorMatrix)

            if totalMaximumError < maximumError:
                totalMaximumError = maximumError

        if totalMaximumError <= tolerance:
            converged = True


    ### CONSOLE PRINT ###

    # Max Iteration Print
    if iteration >= maxIterations:
        print ('Iteration limit reached without convergence')
        print (' ')

    if converged:
        print ('Converged after ' + str(iteration) + ' iterations')
        print (' ')

    if iteration >= maxIterations or converged:
        print ('Bus Voltages')
        print (' ')
        print ('Bus\tPhase A\t\t\tPhase B\t\t\tPhase C\t\t\t')
        print (outputPolarStr)
        print (' ')
        print (' ')
        print ('----- VIOLATIONS -----')
        print (' ')
        print ('Upper Limit : ' + str(upperLimitViolation) + '%')
        print (' ')
        print ('Bus\tPhase A\t\t\tPhase B\t\t\tPhase C\t\t\t')
        for bus in upperViolationArray:
            print (str(bus[0]) + '\t' + "\t\t".join(bus[1]))
        if len(upperViolationArray) == 0:
            print('\t\t\t*** No Violations ***')
        print (' ')
        print (' ')
        print ('Lower Limit : ' + str(lowerLimitViolation) + '%')
        print (' ')
        print ('Bus\tPhase A\t\t\tPhase B\t\t\tPhase C\t\t\t')
        for bus in lowerViolationArray:
            print (str(bus[0]) + '\t' + "\t\t".join(bus[1]))
        if len(lowerViolationArray) == 0:
            print('\t\t\t*** No Violations ***')
        print (' ')
        
    # Next iteration
    iteration = iteration + 1
    calculatedVoltagesOfPreviousIteration = calculatedVoltages

# Finish Algorithm
endTime = time.time()
print ('Calculation Completed! Elapsed Time : ', endTime - startTime)